import csv
import openpyxl

from tqdm import tqdm


def transform(input_file: str, output_file: str):
    """
    Converts the first sheet of a large XLSX file to a CSV file 
    with a progress bar, using a memory-efficient streaming approach.
    """

    # Get the total number of rows for the progress bar.
    total_rows = get_row_count(input_file)
    if total_rows == 0:
        print("Aborting conversion as the sheet has no rows.")
        return

    # Open the workbook and select the first sheet.
    workbook = openpyxl.load_workbook(input_file, read_only=True)
    worksheet = workbook.worksheets[0]

    # Open the output CSV file for writing.
    # The `newline=''` argument is crucial to prevent extra blank rows.
    with open(output_file, "w", newline="", encoding="utf-8") as csvfile:
        csv_writer = csv.writer(csvfile)

        # Initialize progress bar.
        with tqdm(total=total_rows, desc="Converting XLSX to CSV", unit="rows") as pbar:
            # worksheet.iter_rows() is a generator that reads rows on the fly.
            for row in worksheet.iter_rows():
                # Extract the value from each cell. Empty cells become empty strings.
                row_values = [cell.value if cell.value is not None else "" for cell in row]

                # Write the list of values as a new row in the CSV.
                csv_writer.writerow(row_values)

                # Update the progress bar.
                pbar.update(1)

    # Close the workbook to release the file handle.
    workbook.close()


def get_row_count(file_path: str) -> int:
    """
    Efficiently gets the total number of rows from the first sheet
    of an XLSX file without loading the whole file into memory.
    """
    try:
        # Open the workbook in read-only mode for efficiency.
        workbook = openpyxl.load_workbook(file_path, read_only=True)
        # Access the first sheet by its index [0].
        worksheet = workbook.worksheets[0]
        count = worksheet.max_row
        workbook.close()
        return count
    except FileNotFoundError:
        print(f"Error: The file '{file_path}' was not found.")
        return 0
